#importing the necessary libraries
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import itertools
import warnings
warnings.filterwarnings('ignore')

#command to save the data into a worksheet
wb2 = load_workbook('nakuru_links.xlsx')
worksheet = wb2['Sheet']['A']
worksheet1 = wb2['Sheet']['B']
wb = load_workbook('nakuru_hotels.xlsx')
worksheet2 = wb.active
for row, row1 in list(zip(worksheet, worksheet1)):
   # to get the webpage
   page_url =row1.value
   driver = webdriver.Chrome(ChromeDriverManager().install())
   driver.get(page_url)

   #changed from find_elements to find_element
   location = driver.find_element(By.CLASS_NAME, "hp_address_subtitle").text

   #Get list of first row in every room type(notice if you want list you use find_elements)
   roomtypes=driver.find_elements(By.XPATH, "//td[@rowspan]//parent::tr")

   # Start from the first cell. Rows and columns are zero indexed.

   #:)
   for e in roomtypes:
      rtype = e.find_element(By.CLASS_NAME, "hprt-roomtype-icon-link").text
      price = e.find_element(By.CLASS_NAME, "prco-valign-middle-helper").text
      title= row.value
      worksheet2.append([title,rtype,location, price])
   wb.save("nakuru_hotels.xlsx")